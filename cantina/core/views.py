import csv
import io
import json
from datetime import datetime, timedelta
from decimal import Decimal, InvalidOperation
from urllib.parse import urlencode

from django.contrib import messages
from django.contrib.auth import login, logout
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib.auth.forms import AuthenticationForm
from django.db import transaction
from django.db.models import Avg, Count, Prefetch, Q, Sum
from django.http import HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse
from django.utils import timezone
from django.views.decorators.http import require_POST

from .models import Categoria, Cliente, ItemVenda, MovimentacaoEstoque, Produto, Venda

MAX_DESCONTO_PERCENTUAL = Decimal('50.00')
FORMAS_PAGAMENTO_VALIDAS = {codigo for codigo, _ in Venda.FORMA_PAGAMENTO_CHOICES}


def admin_required(view_func):
    return user_passes_test(lambda u: u.is_superuser)(view_func)


def login_view(request):
    if request.method == 'POST':
        form = AuthenticationForm(data=request.POST)
        if form.is_valid():
            user = form.get_user()
            login(request, user)
            return redirect('pos')
    else:
        form = AuthenticationForm()

    return render(request, 'login.html', {'form': form})


@login_required
def logout_view(request):
    logout(request)
    return redirect('login')


@login_required
def pos_view(request):
    categorias = Categoria.objects.filter(ativo=True).prefetch_related('produtos')
    produtos = Produto.objects.filter(ativo=True).select_related('categoria')

    context = {
        'categorias': categorias,
        'produtos': produtos,
    }
    return render(request, 'pos.html', context)


@login_required
@require_POST
def buscar_cliente(request):
    termo = request.POST.get("termo", "").strip()

    if not termo:
        return JsonResponse(
            {"success": False, "error": "Informe o nome ou código do cartão"},
            status=400
        )

    clientes = (
        Cliente.objects
        .filter(ativo=True)
        .filter(
            Q(codigo_cartao__iexact=termo) |
            Q(nome__icontains=termo)
        )
        .order_by("nome")[:10]
    )

    if not clientes.exists():
        return JsonResponse(
            {"success": False, "error": "Cliente não encontrado"},
            status=404
        )

    return JsonResponse({
        "success": True,
        "clientes": [
            {
                "id": c.id,
                "nome": c.nome,
                "codigo_cartao": c.codigo_cartao,
            }
            for c in clientes
        ]
    })


@login_required
@require_POST
def finalizar_venda(request):
    """Finaliza a venda com validações de entrada, desconto e estoque."""
    try:
        data = json.loads(request.body)
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'error': 'JSON inválido'}, status=400)

    cliente_id = data.get('cliente_id')
    itens = data.get('itens', [])
    forma = data.get('forma_pagamento')

    if not itens:
        return JsonResponse({'success': False, 'error': 'Dados incompletos'}, status=400)

    if forma not in FORMAS_PAGAMENTO_VALIDAS:
        return JsonResponse({'success': False, 'error': 'Forma de pagamento inválida'}, status=400)

    try:
        desconto_percentual = Decimal(str(data.get('desconto_percentual', 0)))
    except (InvalidOperation, TypeError, ValueError):
        return JsonResponse({'success': False, 'error': 'Desconto inválido'}, status=400)

    if desconto_percentual < 0 or desconto_percentual > MAX_DESCONTO_PERCENTUAL:
        return JsonResponse({'success': False, 'error': 'Desconto deve estar entre 0% e 50%'}, status=400)

    cliente = None
    if cliente_id:
        cliente = get_object_or_404(Cliente, id=cliente_id, ativo=True)

    if forma == 'FIA' and not cliente:
        return JsonResponse({'success': False, 'error': 'Fiado só é permitido para cliente identificado'}, status=400)

    try:
        with transaction.atomic():
            subtotal = Decimal('0.00')
            itens_validados = []

            for item in itens:
                produto_id = item.get('id')
                quantidade_raw = item.get('quantity')

                if produto_id is None or quantidade_raw is None:
                    return JsonResponse({'success': False, 'error': 'Item inválido na venda'}, status=400)

                try:
                    quantidade = int(quantidade_raw)
                except (TypeError, ValueError):
                    return JsonResponse({'success': False, 'error': 'Quantidade inválida'}, status=400)

                if quantidade <= 0:
                    return JsonResponse({'success': False, 'error': 'Quantidade deve ser maior que zero'}, status=400)

                produto = get_object_or_404(Produto.objects.select_for_update(), id=produto_id, ativo=True)

                # Determine which product's stock is consumed and by how much.
                consumo = quantidade * produto.fator_estoque
                if produto.produto_estoque_id:
                    stock_prod = Produto.objects.select_for_update().get(id=produto.produto_estoque_id)
                else:
                    stock_prod = produto

                if stock_prod.estoque > 0 and consumo > stock_prod.estoque:
                    return JsonResponse(
                        {'success': False, 'error': f'Estoque insuficiente para {produto.nome}'},
                        status=400
                    )

                preco = Decimal(str(produto.preco))
                subtotal_item = preco * quantidade
                subtotal += subtotal_item

                itens_validados.append({
                    'produto': produto,
                    'stock_prod': stock_prod,
                    'consumo': consumo,
                    'quantidade': quantidade,
                    'preco_unitario': preco,
                    'subtotal': subtotal_item,
                })

            desconto_valor = (subtotal * desconto_percentual) / Decimal('100')
            total = subtotal - desconto_valor

            venda = Venda.objects.create(
                cliente=cliente,
                operador=request.user,
                subtotal=subtotal,
                desconto_percentual=desconto_percentual,
                desconto_valor=desconto_valor,
                total=total,
                forma_pagamento=forma,
                paga=(forma != 'FIA'),
                quitada_em=None if forma == 'FIA' else timezone.now(),
            )

            for item in itens_validados:
                produto    = item['produto']
                stock_prod = item['stock_prod']

                if stock_prod.estoque > 0:
                    stock_prod.estoque -= item['consumo']
                    stock_prod.save(update_fields=['estoque'])

                ItemVenda.objects.create(
                    venda=venda,
                    produto=produto,
                    quantidade=item['quantidade'],
                    preco_unitario=item['preco_unitario'],
                    subtotal=item['subtotal'],
                )

            return JsonResponse({
                'success': True,
                'venda_id': venda.id,
                'subtotal': float(subtotal),
                'desconto_percentual': float(desconto_percentual),
                'desconto_valor': float(desconto_valor),
                'total': float(total),
                'message': f'Venda #{venda.id} finalizada com sucesso!',
            })

    except Exception:
        return JsonResponse({'success': False, 'error': 'Erro interno ao finalizar venda'}, status=500)


@login_required
@admin_required
def produtos_list(request):
    termo = request.GET.get('q', '').strip()
    categoria_slug = request.GET.get('categoria', '').strip()

    produtos = Produto.objects.filter(ativo=True).select_related('categoria')

    if termo:
        produtos = produtos.filter(nome__icontains=termo)

    if categoria_slug:
        produtos = produtos.filter(categoria__slug=categoria_slug)

    categorias = Categoria.objects.filter(ativo=True).prefetch_related('produtos')

    vendas = (
        Venda.objects
        .select_related('cliente', 'operador')
        .order_by('-data_hora')[:10]
    )

    vendas_stats = Venda.objects.aggregate(
        qtd_vendas=Count('id'),
        faturamento=Sum('total')
    )

    return render(
        request,
        'produtos.html',
        {
            'produtos': produtos,
            'categorias': categorias,
            'vendas': vendas,
            'vendas_stats': vendas_stats,
            'filtro_q': termo,
            'filtro_categoria': categoria_slug,
        }
    )


MESES_NOMES = [
    '', 'Janeiro', 'Fevereiro', 'Março', 'Abril', 'Maio', 'Junho',
    'Julho', 'Agosto', 'Setembro', 'Outubro', 'Novembro', 'Dezembro',
]


@login_required
@admin_required
def vendas_dashboard(request):
    hoje = timezone.now()

    try:
        mes = int(request.GET.get('mes', hoje.month))
        ano = int(request.GET.get('ano', hoje.year))
        if not (1 <= mes <= 12):
            raise ValueError
    except (ValueError, TypeError):
        mes = hoje.month
        ano = hoje.year

    vendas_mes = Venda.objects.filter(data_hora__year=ano, data_hora__month=mes)
    fiados_mes = vendas_mes.filter(forma_pagamento='FIA')

    total_a_receber_mes = (
        fiados_mes.filter(paga=False).aggregate(total=Sum('total'))['total'] or 0
    )
    faturamento_mes = vendas_mes.aggregate(total=Sum('total'))['total'] or 0
    qtd_vendas_mes = vendas_mes.count()
    ticket_medio_mes = vendas_mes.aggregate(media=Avg('total'))['media'] or 0

    fiados_raw = (
        fiados_mes
        .values('cliente_id', 'cliente__nome')
        .annotate(
            total_fiado=Sum('total'),
            total_pendente=Sum('total', filter=Q(paga=False)),
            qtd_fiado=Count('id'),
            qtd_pendente=Count('id', filter=Q(paga=False)),
        )
        .order_by('cliente__nome')
    )

    fiados_por_cliente = []
    for row in fiados_raw:
        fiados_por_cliente.append({
            'cliente_id': row['cliente_id'],
            'cliente_nome': row['cliente__nome'] or 'Consumidor final',
            'total_fiado': row['total_fiado'] or Decimal('0'),
            'total_pendente': row['total_pendente'] or Decimal('0'),
            'qtd_fiado': row['qtd_fiado'] or 0,
            'qtd_pendente': row['qtd_pendente'] or 0,
            'quitada': (row['qtd_pendente'] or 0) == 0,
        })

    vendas = (
        vendas_mes
        .select_related('cliente', 'operador')
        .order_by('-data_hora')[:50]
    )

    prev_mes = mes - 1 if mes > 1 else 12
    prev_ano = ano if mes > 1 else ano - 1
    next_mes = mes + 1 if mes < 12 else 1
    next_ano = ano if mes < 12 else ano + 1

    return render(request, 'vendas.html', {
        'total_a_receber_mes': total_a_receber_mes,
        'faturamento_mes': faturamento_mes,
        'qtd_vendas_mes': qtd_vendas_mes,
        'ticket_medio_mes': ticket_medio_mes,
        'fiados_por_cliente': fiados_por_cliente,
        'vendas': vendas,
        'mes': mes,
        'ano': ano,
        'mes_nome': MESES_NOMES[mes],
        'prev_mes': prev_mes,
        'prev_ano': prev_ano,
        'next_mes': next_mes,
        'next_ano': next_ano,
    })


@login_required
@admin_required
@require_POST
def quitar_venda(request, venda_id):
    venda = get_object_or_404(Venda, id=venda_id)

    if venda.paga:
        messages.info(request, f'Venda #{venda.id} já está quitada.')
        return redirect('vendas')

    venda.paga = True
    venda.quitada_em = timezone.now()
    venda.save(update_fields=['paga', 'quitada_em'])

    messages.success(request, f'Venda #{venda.id} quitada com sucesso.')
    return redirect('vendas')


@login_required
@admin_required
def exportar_vendas_csv(request):
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="vendas.csv"'

    writer = csv.writer(response)
    writer.writerow([
        'ID', 'Data', 'Cliente', 'Operador', 'Forma de pagamento',
        'Subtotal', 'Desconto (%)', 'Desconto (R$)', 'Total', 'Status'
    ])

    for venda in Venda.objects.select_related('cliente', 'operador').order_by('-data_hora'):
        writer.writerow([
            venda.id,
            venda.data_hora.strftime('%d/%m/%Y %H:%M'),
            venda.cliente.nome if venda.cliente else 'Consumidor final',
            venda.operador.username if venda.operador else '-',
            venda.get_forma_pagamento_display().replace('Cartão', 'Cartao'),
            venda.subtotal,
            venda.desconto_percentual,
            venda.desconto_valor,
            venda.total,
            'Paga' if venda.paga else 'Fiado',
        ])

    return response


@login_required
@admin_required
def exportar_vendas_clientes_csv(request):
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="vendas_por_cliente.csv"'

    writer = csv.writer(response)
    writer.writerow([
        'Cliente', 'Saldo Devedor', 'Qtd Vendas', 'Qtd Fiado'
    ])

    vendas_por_cliente = (
        Venda.objects
        .values('cliente__nome')
        .annotate(
            total_geral=Sum('total'),
            total_fiado=Sum('total', filter=Q(paga=False)),
            qtd=Count('id'),
            qtd_fiado=Count('id', filter=Q(paga=False)),
        )
        .order_by('-total_geral')
    )

    for row in vendas_por_cliente:
        writer.writerow([
            row['cliente__nome'] or 'Consumidor final',
            row['total_fiado'] or Decimal('0'),
            row['qtd'] or 0,
            row['qtd_fiado'] or 0,
        ])

    return response


@login_required
@admin_required
@require_POST
def quitar_cliente_fiados(request, cliente_id):
    cliente = get_object_or_404(Cliente, id=cliente_id, ativo=True)

    try:
        mes = int(request.POST.get('mes', 0))
        ano = int(request.POST.get('ano', 0))
        if not (1 <= mes <= 12) or ano < 2000:
            raise ValueError
    except (ValueError, TypeError):
        mes = ano = 0

    pendentes = Venda.objects.filter(cliente=cliente, paga=False, forma_pagamento='FIA')
    if mes and ano:
        pendentes = pendentes.filter(data_hora__year=ano, data_hora__month=mes)

    qtd = pendentes.count()
    if qtd == 0:
        messages.info(request, f'Nenhuma venda fiada pendente para {cliente.nome}.')
    else:
        pendentes.update(paga=True, quitada_em=timezone.now())
        messages.success(request, f'{qtd} venda(s) de {cliente.nome} foram quitadas.')

    qs = urlencode({'mes': mes, 'ano': ano}) if mes and ano else ''
    return redirect(f"{reverse('vendas')}?{qs}" if qs else reverse('vendas'))


@login_required
@admin_required
def estoque_view(request):
    if request.method == 'POST':
        produto_id = request.POST.get('produto_id')
        tipo = request.POST.get('tipo')
        motivo = request.POST.get('motivo', '').strip()

        try:
            quantidade = int(request.POST.get('quantidade', '0'))
        except ValueError:
            messages.error(request, 'Quantidade inválida.')
            return redirect('estoque')

        if tipo not in {'ENT', 'PER'}:
            messages.error(request, 'Tipo de movimentação inválido.')
            return redirect('estoque')

        if quantidade <= 0:
            messages.error(request, 'A quantidade deve ser maior que zero.')
            return redirect('estoque')

        custo_unitario = None
        if tipo == 'ENT':
            raw_custo = request.POST.get('custo_unitario', '').strip()
            if raw_custo:
                try:
                    custo_unitario = Decimal(raw_custo.replace(',', '.'))
                    if custo_unitario < 0:
                        raise ValueError
                except (InvalidOperation, ValueError):
                    messages.error(request, 'Custo unitário inválido.')
                    return redirect('estoque')

        produto = get_object_or_404(Produto, id=produto_id)

        with transaction.atomic():
            produto = Produto.objects.select_for_update().get(id=produto.id)

            if tipo == 'PER' and quantidade > produto.estoque:
                messages.error(request, f'Estoque insuficiente para perda de {produto.nome}.')
                return redirect('estoque')

            update_fields = ['estoque']

            if tipo == 'ENT':
                if custo_unitario is not None:
                    estoque_atual = Decimal(produto.estoque)
                    custo_atual = Decimal(produto.custo) if produto.custo else Decimal('0')
                    novo_total = estoque_atual + Decimal(quantidade)
                    if novo_total > 0:
                        produto.custo = (
                            (estoque_atual * custo_atual + Decimal(quantidade) * custo_unitario)
                            / novo_total
                        ).quantize(Decimal('0.01'))
                        update_fields.append('custo')
                produto.estoque += quantidade
            else:
                produto.estoque -= quantidade

            produto.save(update_fields=update_fields)

            MovimentacaoEstoque.objects.create(
                produto=produto,
                tipo=tipo,
                quantidade=quantidade,
                custo_unitario=custo_unitario,
                motivo=motivo,
                usuario=request.user,
            )

        messages.success(request, 'Movimentação de estoque registrada com sucesso.')
        return redirect('estoque')

    produtos = Produto.objects.filter(ativo=True).select_related('categoria').order_by('nome')
    movimentacoes = MovimentacaoEstoque.objects.select_related('produto', 'usuario')[:40]

    return render(
        request,
        'estoque.html',
        {
            'produtos': produtos,
            'movimentacoes': movimentacoes,
        }
    )


@login_required
@admin_required
def lancar_venda_mensal(request):
    """Lançamento manual de venda mensal (sem desconto de estoque)."""
    FORMAS_VALIDAS = {'DIN', 'CAR', 'PIX'}

    if request.method == 'POST':
        forma = request.POST.get('forma_pagamento', '').strip()
        data_str = request.POST.get('data', '').strip()

        if forma not in FORMAS_VALIDAS:
            messages.error(request, 'Forma de pagamento inválida.')
            return redirect('lancamento_mensal')

        try:
            data_venda = datetime.strptime(data_str, '%Y-%m-%d').date()
        except (ValueError, TypeError):
            messages.error(request, 'Data inválida.')
            return redirect('lancamento_mensal')

        produtos_qs = Produto.objects.filter(ativo=True)
        itens = []
        for produto in produtos_qs:
            qty_str = request.POST.get(f'produto_{produto.id}', '0').strip() or '0'
            try:
                qty = int(qty_str)
            except ValueError:
                qty = 0
            if qty > 0:
                itens.append({'produto': produto, 'quantidade': qty})

        if not itens:
            messages.error(request, 'Adicione pelo menos um produto com quantidade maior que zero.')
            return redirect('lancamento_mensal')

        with transaction.atomic():
            subtotal = sum(
                Decimal(str(i['produto'].preco)) * i['quantidade']
                for i in itens
            )
            data_hora = timezone.make_aware(
                datetime(data_venda.year, data_venda.month, data_venda.day, 12, 0, 0)
            )

            venda = Venda.objects.create(
                cliente=None,
                operador=request.user,
                data_hora=data_hora,
                subtotal=subtotal,
                desconto_percentual=Decimal('0'),
                desconto_valor=Decimal('0'),
                total=subtotal,
                forma_pagamento=forma,
                paga=True,
                quitada_em=data_hora,
                observacao='Lançamento mensal manual',
            )

            for item in itens:
                preco = Decimal(str(item['produto'].preco))
                ItemVenda.objects.create(
                    venda=venda,
                    produto=item['produto'],
                    quantidade=item['quantidade'],
                    preco_unitario=preco,
                    subtotal=preco * item['quantidade'],
                )

        messages.success(request, f'Venda #{venda.id} lançada com sucesso! Total: R$ {subtotal:.2f}')
        return redirect('lancamento_mensal')

    categorias = (
        Categoria.objects
        .filter(ativo=True)
        .prefetch_related(
            Prefetch('produtos', queryset=Produto.objects.filter(ativo=True).order_by('nome'), to_attr='produtos_ativos')
        )
        .order_by('ordem', 'nome')
    )
    hoje = timezone.localdate().isoformat()

    return render(request, 'lancamento.html', {
        'categorias': categorias,
        'hoje': hoje,
    })


SLUG_SERVICOS = 'servicos'


def _build_relatorio_rows(ano, mes):
    """Returns (rows, totals) for the monthly report. Shared by dashboard and XLSX views."""
    itens = list(
        ItemVenda.objects
        .filter(venda__data_hora__year=ano, venda__data_hora__month=mes)
        .values('produto__id', 'produto__nome', 'produto__categoria__slug', 'produto__categoria__nome')
        .annotate(qtd_total=Sum('quantidade'), valor_total=Sum('subtotal'))
        .order_by('produto__categoria__nome', 'produto__nome')
    )

    produto_ids = [i['produto__id'] for i in itens]
    custos = {p.id: p.custo for p in Produto.objects.filter(id__in=produto_ids)}

    rows = []
    cantina_valor = Decimal('0')
    cantina_custo = Decimal('0')
    servicos_valor = Decimal('0')
    servicos_custo = Decimal('0')
    total_qtd = 0

    for item in itens:
        qtd = item['qtd_total'] or 0
        valor_total = Decimal(str(item['valor_total'] or 0))
        custo_unit = Decimal(str(custos.get(item['produto__id'], 0) or 0))
        valor_unit = (valor_total / qtd).quantize(Decimal('0.01')) if qtd else Decimal('0')
        custo_total = (custo_unit * qtd).quantize(Decimal('0.01'))
        lucro = (valor_total - custo_total).quantize(Decimal('0.01'))
        is_servico = item['produto__categoria__slug'] == SLUG_SERVICOS

        rows.append({
            'nome': item['produto__nome'],
            'categoria': item['produto__categoria__nome'],
            'valor_unit': valor_unit,
            'custo_unit': custo_unit,
            'qtd': qtd,
            'valor_total': valor_total,
            'custo_total': custo_total,
            'lucro': lucro,
            'is_servico': is_servico,
        })

        total_qtd += qtd
        if is_servico:
            servicos_valor += valor_total
            servicos_custo += custo_total
        else:
            cantina_valor += valor_total
            cantina_custo += custo_total

    totals = {
        'cantina_valor': cantina_valor,
        'cantina_custo': cantina_custo,
        'cantina_lucro': cantina_valor - cantina_custo,
        'servicos_valor': servicos_valor,
        'servicos_custo': servicos_custo,
        'servicos_lucro': servicos_valor - servicos_custo,
        'geral_valor': cantina_valor + servicos_valor,
        'geral_custo': cantina_custo + servicos_custo,
        'geral_lucro': (cantina_valor + servicos_valor) - (cantina_custo + servicos_custo),
        'total_qtd': total_qtd,
    }
    return rows, totals


@login_required
@admin_required
def relatorio_mensal_dashboard(request):
    now = timezone.now()
    try:
        mes = int(request.GET.get('mes', now.month))
        ano = int(request.GET.get('ano', now.year))
        if not (1 <= mes <= 12):
            raise ValueError
    except (ValueError, TypeError):
        mes = now.month
        ano = now.year

    rows, totals = _build_relatorio_rows(ano, mes)

    prev_mes = mes - 1 if mes > 1 else 12
    prev_ano = ano if mes > 1 else ano - 1
    next_mes = mes + 1 if mes < 12 else 1
    next_ano = ano if mes < 12 else ano + 1

    return render(request, 'relatorio.html', {
        'rows': rows,
        'totals': totals,
        'mes': mes,
        'ano': ano,
        'mes_nome': MESES_NOMES[mes],
        'meses_list': list(enumerate(MESES_NOMES))[1:],
        'prev_mes': prev_mes,
        'prev_ano': prev_ano,
        'next_mes': next_mes,
        'next_ano': next_ano,
    })


@login_required
@admin_required
def relatorio_mensal_xlsx(request):
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    now = timezone.now()
    try:
        mes = int(request.GET.get('mes', now.month))
        ano = int(request.GET.get('ano', now.year))
        if not (1 <= mes <= 12):
            raise ValueError
    except (ValueError, TypeError):
        mes = now.month
        ano = now.year

    rows, totals = _build_relatorio_rows(ano, mes)

    wb = Workbook()
    ws = wb.active
    ws.title = f"{mes:02d}-{ano}"

    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='1F2937', end_color='1F2937', fill_type='solid')
    center = Alignment(horizontal='center')
    currency_fmt = 'R$ #,##0.00'
    total_font = Font(bold=True)

    headers = ['Categoria', 'Produto', 'Valor Unitário', 'Custo', 'Qtd', 'Valor Total', 'Custo Total', 'Lucro']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center

    for row_num, r in enumerate(rows, 2):
        ws.cell(row=row_num, column=1, value=r['categoria'])
        ws.cell(row=row_num, column=2, value=r['nome'])
        ws.cell(row=row_num, column=3, value=float(r['valor_unit'])).number_format = currency_fmt
        ws.cell(row=row_num, column=4, value=float(r['custo_unit'])).number_format = currency_fmt
        ws.cell(row=row_num, column=5, value=r['qtd'])
        ws.cell(row=row_num, column=6, value=float(r['valor_total'])).number_format = currency_fmt
        ws.cell(row=row_num, column=7, value=float(r['custo_total'])).number_format = currency_fmt
        ws.cell(row=row_num, column=8, value=float(r['lucro'])).number_format = currency_fmt

    # Two summary rows
    cantina_row = len(rows) + 2
    geral_row = cantina_row + 1

    cantina_fill = PatternFill(start_color='DBEAFE', end_color='DBEAFE', fill_type='solid')
    geral_fill = PatternFill(start_color='D1FAE5', end_color='D1FAE5', fill_type='solid')

    for col, (label, fill, key_val, key_custo, key_lucro) in enumerate([
        ('Total Cantina', cantina_fill, 'cantina_valor', 'cantina_custo', 'cantina_lucro'),
        ('Total Geral',   geral_fill,  'geral_valor',   'geral_custo',   'geral_lucro'),
    ], 0):
        row = cantina_row + col
        fill_obj = cantina_fill if col == 0 else geral_fill
        label_text = 'Total Cantina' if col == 0 else 'Total Geral (c/ Serviços)'
        ws.cell(row=row, column=1, value=label_text).font = total_font
        ws.cell(row=row, column=6, value=float(totals[key_val])).number_format = currency_fmt
        ws.cell(row=row, column=7, value=float(totals[key_custo])).number_format = currency_fmt
        ws.cell(row=row, column=8, value=float(totals[key_lucro])).number_format = currency_fmt
        for c in range(1, 9):
            cell = ws.cell(row=row, column=c)
            cell.font = total_font
            cell.fill = fill_obj

    col_widths = [18, 28, 16, 12, 8, 16, 14, 14]
    for col, width in enumerate(col_widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = width

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"relatorio_{ano}_{mes:02d}.xlsx"
    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


@login_required
@admin_required
def baixar_fatura_cliente(request, cliente_id, ano, mes):
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    cliente = get_object_or_404(Cliente, id=cliente_id)

    vendas = (
        Venda.objects
        .filter(cliente=cliente, forma_pagamento='FIA', data_hora__year=ano, data_hora__month=mes)
        .prefetch_related('itens__produto')
        .order_by('data_hora')
    )

    wb = Workbook()
    ws = wb.active
    ws.title = f"Fatura {mes:02d}-{ano}"

    ws.merge_cells('A1:F1')
    title = ws['A1']
    title.value = f"Fatura — {cliente.nome} — {MESES_NOMES[mes]}/{ano}"
    title.font = Font(bold=True, size=13)

    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='1F2937', end_color='1F2937', fill_type='solid')
    currency_fmt = 'R$ #,##0.00'

    for col, h in enumerate(['Data', 'Produto', 'Qtd', 'Preço Unit.', 'Total', 'Status'], 1):
        cell = ws.cell(row=2, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    row_num = 3
    total_geral = Decimal('0')
    total_pendente = Decimal('0')

    for venda in vendas:
        for item in venda.itens.all():
            ws.cell(row=row_num, column=1, value=venda.data_hora.strftime('%d/%m/%Y'))
            ws.cell(row=row_num, column=2, value=item.produto.nome)
            ws.cell(row=row_num, column=3, value=item.quantidade)
            ws.cell(row=row_num, column=4, value=float(item.preco_unitario)).number_format = currency_fmt
            ws.cell(row=row_num, column=5, value=float(item.subtotal)).number_format = currency_fmt
            ws.cell(row=row_num, column=6, value='Pago' if venda.paga else 'Pendente')
            row_num += 1
        total_geral += venda.total
        if not venda.paga:
            total_pendente += venda.total

    total_font = Font(bold=True)
    total_fill = PatternFill(start_color='E5E7EB', end_color='E5E7EB', fill_type='solid')
    pending_font = Font(bold=True, color='DC2626')
    pending_fill = PatternFill(start_color='FEE2E2', end_color='FEE2E2', fill_type='solid')

    for c in range(1, 7):
        ws.cell(row=row_num, column=c).fill = total_fill
    ws.cell(row=row_num, column=1, value='TOTAL').font = total_font
    ws.cell(row=row_num, column=5, value=float(total_geral)).number_format = currency_fmt
    ws.cell(row=row_num, column=5).font = total_font

    if total_pendente > 0:
        row_num += 1
        for c in range(1, 7):
            ws.cell(row=row_num, column=c).fill = pending_fill
        ws.cell(row=row_num, column=1, value='PENDENTE').font = pending_font
        ws.cell(row=row_num, column=5, value=float(total_pendente)).number_format = currency_fmt
        ws.cell(row=row_num, column=5).font = pending_font

    for col, width in enumerate([12, 30, 6, 14, 14, 10], 1):
        ws.column_dimensions[ws.cell(row=2, column=col).column_letter].width = width

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    safe_nome = ''.join(c if c.isalnum() else '_' for c in cliente.nome)
    filename = f"fatura_{safe_nome}_{ano}_{mes:02d}.xlsx"
    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response
